import cv2 as cv
import numpy as np
import math
from openpyxl import load_workbook



#                                               SOURCES
# Utilizing Yolov3, https://pjreddie.com/darknet/yolo/
# Utilzing OpenCV
# 
#
# Hello, before runing the program make sure to clear the files cars.xlxs and vel.xlxs
#
# The current code is setup with an "acceptace window" to remove this change the following lines...
#Line 89
#Line 114 and 115  <----- just a visual representation of the acceptance window
#and remove the if statement from Line 328

#Loading Yolo
def loadYolo():
    net = cv.dnn.readNet("yolov3.weights", "yolov3.cfg")
    classes = []
    with open("coco.names", "r") as f:  
        classes = [line.strip() for line in f.readlines()]
    layer_names = net.getLayerNames()
    output_layers = [layer_names[i[0] - 1] for i in net.getUnconnectedOutLayers()]
    return net, classes , output_layers 

#Yolo Detenction
def detection(img, net ,Output_Layer):
    blob = cv.dnn.blobFromImage(img, 0.00392, (608,608), (0,0,0), True, crop=False)
    net.setInput(blob)
    outs = net.forward(Output_Layer)
    return blob,outs

def process(outs, height, width):
    accept = ["car","truck","bus"] #Last change removed bike and train because they are not really "Cars"
    class_ids = []
    confidences = []
    boxes = []
    framedata = []
    data = []
    for out in outs:
        for detection in out:
            scores = detection[5:]
            class_id = np.argmax(scores)
            check = str(classes[class_id])

            #since we are utilizing the coco dataset, we dont't want other things other than "vehicles", so they are filter out here.
            if check in accept:
                confidence = scores[class_id]
                if confidence > 0.4:
                     #Object detected
                     center_x = int(detection[0] * width)
                     center_y = int(detection[1] *height)
                     w = int(detection[2] * width)
                     h = int(detection[3] * height)
                     x = int(center_x - w/ 2)
                     y = int(center_y - h / 2)

                     #Declare an ending area for the video, in case that the camera is setup in such a way that the yolo detection system cannot detect cars properly past a certain region.
                     #if y > 450:
                     #    break

                     
                     #halfy,halfx provides the center of the bounding box of each car.
                     halfy = int(y+(h/2))
                     halfx = int(x+(w/2))   
                     data = [halfx,halfy]
                     framedata.append(data)
                     ##########print("center of vehicle is: " , halfx ,  " " , halfy )


                     boxes.append([x,y,w,h])
                     confidences.append(round(float(confidence),3))
                     class_ids.append(class_id)
    return class_ids, confidences, boxes, framedata



# apply non-max suppression on the list of cars that will be used for tracking.
def supression(boxes,confidences, framedata):
    indexes = cv.dnn.NMSBoxes(boxes, confidences, 0.4, 0.3)
    cars = []

    for i in range(len(framedata)):
        if i in indexes:
            #change to change the acceptance window of new vehicles.
            if framedata[i][1] > 50:
                cars.append(framedata[i])
                
    return cars

            

    



def draw(img,class_ids,confidences,boxes,width,carpath,velocity):
    
    indexes = cv.dnn.NMSBoxes(boxes, confidences, 0.4, 0.3)
    font = cv.FONT_HERSHEY_PLAIN
    carnumber = -1  #if car was not found in tracker then it will display -1

    
    color1 = (135,225,129)     #Car is not moving/ slow
    color2 = (34,212,247)  #Car is moving
    color3 = (44,44,188)   #Car is moving too fast

    
    #newcar acceptance window
    cv.line(img, pt1=(0,75), pt2=(1000,75), color=(255,0,0), thickness=2)
    cv.line(img, pt1=(0,50), pt2=(1000,50), color=(0,255,0), thickness=2)



    #creates the line for the end border of the detection system
    #cv.line(img, pt1=(0,450), pt2=(1000,450), color=(0,0,255), thickness=2)

    for i in range(len(boxes)):
        if i in indexes:
            x,y,w,h =boxes[i]
            curvel = 0

            halfy = int(y+(h/2))
            halfx = int(x+(w/2))   
            data = [halfx,halfy]
            

            iterator = 0
            for i in carpath:
                if data == i[len(i)-1]: 
                    carnumber = i[0][0]
                    label3 = str(carnumber)

                    
                    frames = len(velocity[iterator][1]) # since carpath and velocity lists are a 1:1 (car position to velocity list)
                    if frames >= 6: 
                        #Take the last 5 frame velocity and return the avg    
                        curvel = (velocity[iterator][1][len(velocity[iterator][1])-1] + velocity[iterator][1][len(velocity[iterator][1])-2] + velocity[iterator][1][len(velocity[iterator][1])-3] + velocity[iterator][1][len(velocity[iterator][1])-4] + velocity[iterator][1][len(velocity[iterator][1])-5]) /5
                        label5 = str(round(curvel,3))
                        cv.putText(img, label5, (x,y-1), font, 1, (0,0,255), 2)
                        #print(velocity[iterator][1][len(velocity[iterator][1])-1])
                        #print(curvel)
                        #input("Press Enter to continue...")
                    elif frames >= 2: 
                        curvel = velocity[iterator][1][len(velocity[iterator][1])-1]
                        label5 = str(curvel)
                        cv.putText(img, label5, (x,y-1), font, 1, (0,0,255), 2)
                    



                    #cv.circle(img, (halfx,halfy), radius= 0, color = (0,255,0), thickness= 5 ) #center of box


                    if curvel < 25: #Car is not moving or the car is moving too slow
                        cv.rectangle(img,(x,y), (x+w,y+h), color1, 2)           #carbox
                    elif curvel < 115: #Car is moving at a moderate speed
                        cv.rectangle(img,(x,y), (x+w,y+h), color2, 2) 
                    else:   #Car is moving too fast
                        cv.rectangle(img,(x,y), (x+w,y+h), color3, 2)

                    cv.putText(img, label3, (halfx,halfy), font, 1, (0,0,255), 2)  #carnumber

                iterator = iterator + 1
        


    cv.imshow("Image", img)
    return img


#Tracker 
#1st instance all cars that were detected in the first frame will become a car being tracked ["id","x","y"], id being the car number.
#
#for every consecutive frame check what car is closest to each individual car, if a car has no match then it will be considered a "problem"
#   problems are counted to keep check of number of cars being droped in each frame.
#
#for a problem ["M","times","id","x","y"] M stands for missing, and 'times' is the amount of frames which the car was missing after a certain amount of frames 
#   specified in the code the car will be dropped.
#   
#after a car is matched it will be deleted from new frame, if there are more cars in newframe but every car in oldframe has been properly matched, then if those frames
#   meets the criteria of possible new car location, then those cars wills be tracked.
#
def tracker(oldframe,newframe,carpath,noc, velocity):
    #framerate of video
    framerate = 30
    time = 2/framerate


    track = carpath
    edit = []       
    best = []     
    bestscore = 1000          
    current = 0
    create = []
    problem = 0
    problemcount  = 0

    if len(oldframe) == 0:
        number = 0  #number of car being tracked
        store = []  #store will be the first sequence of cars that go into the tracking.
        storev = []
        for i in newframe:
            
            cart = [i]
            cart.insert(0,[number])
            store.append(cart)
            

            cartv = [[number],[]]
            storev.append(cartv)

            number = number +1  
            
        return store,newframe,number,storev



   # print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
   # print("OLD    " , oldframe)
    #print("NEW    " , newframe)



    for oldi in oldframe:
        #print(oldframe)
        #print("~~~~~~~~~~~~~~~~~~~~~~")
        #print(oldi)
        #for each car in the old frame check too see who is closest
        for newi in newframe:
            if(oldi[0] != 'M'):

                score = math.sqrt( pow(oldi[0]-newi[0],2) + pow(oldi[1]-newi[1],2))
                #print(score)
            else:
                score = math.sqrt( pow(oldi[2]-newi[0],2) + pow(oldi[3]-newi[1],2))
            if len(best) == 0:
                best = newi
                bestscore = score
            elif score < bestscore:
                best = newi
                bestscore = score

        if bestscore == 1000:
            problem = 1


        #Check if other cars in oldlist have a better score than the current car, if they do then that car is missing.
        if len(best) != 0:
            for other in oldframe:
                if other[0] != 'M':
                    score = math.sqrt( pow(best[0]-other[0],2) + pow(best[1]-other[1],2))
                else:
                    score = math.sqrt( pow(best[0]-other[2],2) + pow(best[1]-other[3],2))
                if score < bestscore:
                    problem = 1
                    problemcount = problemcount + 1 #Keep count of problems in current frame
                    break
        

        #Testing1
        #if the car does not have a problem, but the car is too far print that
        if problem != 1:
            #print("The best score is:" ,bestscore)
            if(bestscore > 180):  #the score is too great for it to be considered a possible point of car.
                print(best)
                print(oldi)
                print("THE CAR IS TOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO FAR")
                print("The best score is:" ,bestscore)
                problem = 1



        #Deal with cars that are missing
        if problem == 1:
            #M for missing
            # ['M',#times,x,y]
            #if a car is not found to match the car that is already MIA
            #then add 1 to #times upto 3 times, then drop the entire track on that vehicle
            if(oldi[0] != 'M'):
                edit = [oldi[0],oldi[1]]        # edit = oldi creates issues.
                edit.insert(0,'M')
                edit.insert(1,1)
                create.append(edit)
                track[current].append(edit)
            else:
                edit = [oldi[0],oldi[1],oldi[2],oldi[3]]
                edit[1] = edit[1]+1
                if edit[1] != 3:
                    create.append(edit)
                track[current].append(edit)

            problem = 0
            

        else:
            index = 0
            for i in newframe:
                if i == best:
                    create.append(i)
                    del newframe[index]
                index = index +1
            #we can track speed here!
            # current = old frame , best = new frame 
            #  time = number of frames (2)/ framerate
            #  velocity = distance / time
            #  distance = bestscore 
            vel = int(bestscore/time)
            velocity[current][1].append(vel)
            #print(velocity)
            track[current].append(best)


        #reset

        best = []
        bestscore = 1000
        current = current + 1

    #deal with leftovers
    print("this frame is not empty ! ! ! ! Number of items left: ", len(newframe), newframe)
    for i in newframe:
        #change to change the acceptance window of new vehicles.
        if i[1] < 75 and i[1] > 50:
            edit = [i[0],i[1]]
            create.append(edit)
            car_add = [i]
            car_add.insert(0,[noc])
            #print(car_add)
            track.append(car_add)

            carveloc = [[noc],[]]
            velocity.append(carveloc)

            noc = noc +1



    print("Problems")
    print(problemcount)
    return track,create,noc,velocity



    #Since the weights are pretrained in small images 
    #Rescaling the image can provide better results for the Yolo Detection 
def rescaleFrame(frame,scale =0.5):
    #  works with Images, Videos, and live video
    width = int(frame.shape[1] * scale)
    height = int(frame.shape[0] * scale)

    dimension = (width,height)

    return cv.resize(frame,dimension,interpolation=cv.INTER_AREA)


def printcars(carpath,velocity):
    #print("_________________________________________________________________________________")
    index = 0 
    for i in carpath:
        for x in i:
            if x[0] == 'M':
                if x[1] == 3:  # if this car has not been found in #frames then drop.

                    wb = load_workbook('cars.xlsx')
                    work_sheet = wb.active # Get active sheet
                    column = 1
                    #since cars begin with 0 and excel begins with 1, add 1 to the row for cell
                    row = carpath[index][0][0] 
                    for car in carpath[index]:
                        if len(car) > 2:
                            work_sheet.cell(row = row+1,column = column).value = str(car[0])+','+str(car[1])+","+str(car[2])+','+str(car[3])
                        elif len(car) > 1:
                            work_sheet.cell(row = row+1,column = column).value = str(car[0])+','+str(car[1])
                        else:
                            work_sheet.cell(row = row+1,column = column).value = row
                        column = column +1
                        #print(i)
                        #print("column = ", column)

                    wb.save('cars.xlsx')

                    wb = load_workbook('vel.xlsx')
                    work_sheet = wb.active # Get active sheet
                    column = 2
                    #since cars begin with 0 and excel begins with 1, add 1 to the row for cell
                    row = carpath[index][0][0] 
                    work_sheet.cell(row = row+1,column = 1).value = row
                    for v in velocity[index][1]:
                        #print(v)
                        work_sheet.cell(row = row+1,column = column).value = str(v)
                        column = column +1
                    
                    wb.save('vel.xlsx')



                        
                    #input("Press Enter to continue...")
                    del carpath[index]
                    del velocity[index]
        index = index +1
    
    #for i in velocity:
    #    print(i[0])
    #print("_____________________________")
    #for i in carpath:
    #    print(i[0])
    #print(len(carpath))
                    







net, classes, Output_Layer = loadYolo()
cap = cv.VideoCapture("cars.mp4")

frame_width = int(cap.get(3) * 0.5)
frame_height = int(cap.get(4) * 0.5)


size = (frame_width,frame_height)
result = cv.VideoWriter('filename.avi',  cv.VideoWriter_fourcc(*'MJPG'), 20, size) 
oldframe = []   #the x,y data for cars from the previous frame
newframe = []   #the x,y data for cars from the newest frame.
carpath = []    #the track or collection of data for each individual vehicle
velocity = []   #velocity of vehicles 
noc = 0 #number of cars tracked so far




while True:
    isTrue, frame = cap.read()
    framer = rescaleFrame(frame)
    height, width, channels = framer.shape
    blob, outs = detection(framer,net, Output_Layer)
    class_ids, confidences, boxes, newframe = process(outs, height, width)
    newframe = supression(boxes, confidences, newframe)
    carpath, oldframe,noc, velocity = tracker(oldframe,newframe,carpath,noc, velocity)
    drawn = draw(framer,class_ids,confidences,boxes,width,carpath,velocity)
    result.write(drawn)

    printcars(carpath,velocity)
    
    if cv.waitKey(5) & 0xFF==ord('d'):
        break
    #uncomment to go frame by frame
    #input("Press Enter to continue...")

print("Done!")
cap.release()
cv.destroyAllWindows()

